import json
import pandas as pd
from openpyxl import load_workbook

def appendWorkBook(dataframe):
    df = dataframe
    book = load_workbook('bleh.xlsx')
    writer = pd.ExcelWriter('bleh.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header= False)

    writer.save()


jsonData = open("fundimentals.json")
raw = json.load(jsonData)
df = pd.json_normalize(raw)
appendWorkBook(df)
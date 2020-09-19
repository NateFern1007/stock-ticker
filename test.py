import requests
import json
import argparse
import pandas as pd
from pandas_datareader import data as pdr
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from ap import Nasdaq
import yaml


def appendWorkBook(dataframe):
    df = dataframe
    book = load_workbook('test.xlsx')
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,
                    header=False)

    writer.save()


with open(r'config.yaml') as file:
    # The FullLoader parameter handles the conversion from YAML
    # scalar values to Python the dictionary format
    cfg = yaml.load(file, Loader=yaml.FullLoader)

    print(cfg["stockType"])

    mutualfunds = cfg["stockType"]["MutualFunds"]
    etfs = cfg["stockType"]["ExchangeTradedFunds"]
    stocks = cfg["stockType"]["CommonStock"]

    for key in cfg["stockType"].keys():
        if key == "MutualFunds":
            q = cfg["stockType"]["MutualFunds"] = "mutualfunds"
            print(q)

    # stockDict = {"symbols": {}}
    # mutualFundsDict = {"mutualfunds": {}}
    # etfsDict = {"etfs": {}}
    # commonStockDict = {"stocks": {}}

    # mutualFundsDict["mutualfunds"] = stock_dict['stockType']["MutualFunds"]
    # etfsDict["etf"] = stock_dict['stockType']["ExchangeTradedFunds"]
    # commonStockDict["stocks"] = stock_dict['stockType']["CommonStock"]

    # assetClass = "".join(stockDict['symbols'].keys())
    # print("AssetClasses", assetClass)


    # print(commonStockDict['stocks'])
    # stockDict['symbols'].update(mutualFundsDict)
    # stockDict['symbols'].update(etfsDict)
    # stockDict['symbols'].update(commonStockDict)
    # print(stockDict)
    # with open('summary-copy.json','w') as fp:
    #     json.dump(stockDict, fp, indent = 4, ensure_ascii=False) # dumps to json file.
    # # symbols = ", ".join([str(value) for elem in commonStockDict.values() for value in elem])

    # for assetClass in stockDict["symbols"]:
    #     print(assetClass)
    #     if assetClass == "stocks":
    #         symbols = commonStockDict["stocks"]
    #         getQuote(symbols)
    #     elif assetClass == "mutualfunds":
    #         symbols = mutualFundsDict["mutualfunds"]
    #         getQuote(symbols)
    #     elif assetClass == "etf":
    #         symbols = etfsDict["etf"]
    #         getQuote(symbols)
    #     else:
    #         print("Error please choose an approperate ticker! ")

# api = Nasdaq("AAPL", "stocks")
# summary = api.getStockInfo("AAPL")


# dropItems(summary)

# dataframe.drop("data_secondaryData")
# dataframe.drop("data_keyStats_Volume_label")
# dataframe.drop("data_keyStats_ExpenseRatio_label")
# dataframe.drop("data_keyStats_AUM_label")
# dataframe.drop("data_keyStats_AverageSpread_label")
# dataframe.drop("data_keyStats_AverageSpread_value")
# dataframe.drop("message")
# dataframe.drop("status_bCodeMessage")
# dataframe.drop("status_developerMessage")



#
# print(api.getMarketStatus())

# with open('summary-copy.json','w') as fp:
#     json.dump(stock_dict, fp, indent = 4, ensure_ascii=False) # dumps to json file.

# ticker1 = "AAPL"
# assetClass1 = "stocks"
#
# api = Nasdaq(ticker1, assetClass1)
# # chart = api.chartDetails()
# print(api.getMarketStatus())


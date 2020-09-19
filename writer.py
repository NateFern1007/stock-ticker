import json
import pandas as pd
from openpyxl import load_workbook

def flatten_dict(dd, separator ='_', prefix =''):
    return { prefix + separator + k if prefix else k : v
             for kk, vv in dd.items()
             for k, v in flatten_dict(vv, separator, kk).items()
             } if isinstance(dd, dict) else { prefix : dd }


def appendWorkBook(dataframe):
    df = dataframe
    book = load_workbook('test.xlsx')
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header= False)

    writer.save()


f = open("out_info.json")
json_data = json.load(f)
data = flatten_dict(json_data)
# for key in data.keys():
#    print(key)


data.pop("info_data_exchange")
data.pop("info_data_isNasdaqListed")
data.pop("info_data_isNasdaq100")
data.pop("info_data_isHeld")
data.pop("info_data_primaryData_isRealTime")
data.pop("info_data_secondaryData_isRealTime")
data.pop("info_data_keyStats_Volume_label")
data.pop("info_data_keyStats_PreviousClose_label")
data.pop("info_data_keyStats_OpenPrice_label")
data.pop("info_data_keyStats_MarketCap_label")
data.pop("info_data_marketStatus")
data.pop("info_data_primaryData_lastTradeTimestamp")
# data.pop("info_data_keyStats_ExpenseRatio_label")
# data.pop("info_data_keyStats_AUM_label")
data.pop("info_data_assetClass")
# data.pop("info_data_keyStats_AverageSpread_label")
# data.pop("info_data_keyStats_AverageSpread_value")
data.pop("info_message")
data.pop("info_status_bCodeMessage")
data.pop("info_status_developerMessage")
data.pop("info_status_rCode")

## summary
data.pop("summary_data_symbol")
data.pop("summary_data_summaryData_Exchange_label")
data.pop("summary_data_summaryData_Sector_label")
data.pop("summary_data_summaryData_Industry_label")
data.pop("summary_data_summaryData_OneYrTarget_label")
data.pop("summary_data_summaryData_TodayHighLow_label")
data.pop("summary_data_summaryData_ShareVolume_label")
data.pop("summary_data_summaryData_AverageVolume_label")
data.pop("summary_data_summaryData_PreviousClose_label")
data.pop("summary_data_summaryData_FiftTwoWeekHighLow_label")
data.pop("summary_data_summaryData_MarketCap_label")
data.pop("summary_data_summaryData_PERatio_label")
data.pop("summary_data_summaryData_ForwardPE1Yr_label")
data.pop("summary_data_summaryData_EarningsPerShare_label")
data.pop("summary_data_summaryData_AnnualizedDividend_label")
data.pop("summary_data_summaryData_DividendPaymentDate_label")
data.pop("summary_data_summaryData_Yield_label")
data.pop("summary_data_summaryData_Beta_label")
data.pop("summary_data_summaryData_ExDividendDate_label")
data.pop("summary_data_assetClass")
data.pop("summary_data_additionalData")
data.pop("summary_message")
data.pop("summary_status_rCode")
data.pop("summary_status_bCodeMessage")
data.pop("summary_status_developerMessage")
print(data, end="\n")

df = pd.json_normalize(data)
appendWorkBook(df)

# dumps JSON File for Testing Purposes:
# with open('out7.json', 'w') as fp:
#     json.dump(data, fp, indent=4, ensure_ascii=False)  # dumps to json file.


